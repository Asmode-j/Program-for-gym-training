import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from tkinter import *

def osnova():
    for i,a in enumerate(days, 1):
        i+=1
        i=str(i)
        sheet['A'+i].font=font
        sheet['A'+i].fill=fill
        sheet['A'+i]= a
    for i in months.keys():
        sheet[i+'1'].font=font
        sheet[i+'1'].fill=fill
        sheet[i+'1'].alignment=alignment
        sheet.column_dimensions[i].width = 13
        sheet[i+'1']=months[i]
    for i,a in zip(grups_a, grups):
        sheet.column_dimensions[i].width = 15
        sheet[i+'1'].font=font
        sheet[i+'1'].alignment=alignment
        sheet[i+'1'].fill=fill_grups
        sheet[i+'1']= a
    for i,d in zip(months.values(), range(2, 25 , 2)):
        sheet.merge_cells(f'N{d}:T{d}')
        sheet[f'N{d}'].font=font
        sheet[f'N{d}'].fill=fill
        sheet[f'N{d}'].alignment=alignment
        sheet[f'N{d}'] = f'{i}'
    

def clicked2():
    a1=txt1.get()
    a2=txt2.get()
    a3=txt3.get()
    a4=txt4.get()
    a5=txt5.get()
    a6=txt6.get()
    a7=txt7.get()
    m=[a1,a2,a3,a4,a5,a6,a7]
    for i,m in zip(res_grups,m):
        file=open('Data.txt' , 'r')
        print(m)
        spisok_dat=file.readlines()
        o=spisok_dat[-1][0]
        dm=data.month
        if int(o) != dm:
            file=open(f'{res_grups[i]}.txt' , 'a')
            file.write('\n')
            file.close()
        file=open(f'{res_grups[i]}.txt' , 'a')
        file.write(f' {m}')
        file.close()
    otmetka_grups()
        

def otmetka():
    for mes_den in vse_data:
        mes_den=mes_den.split()
        mes=int(mes_den[0])+1
        den=str(int(mes_den[1])+1)
        for a,i in enumerate(months.keys(), 2):
            if mes==a:
                sheet[i+den].fill=fill_otmetka

def otmetka_grups():
    for i,ga in zip(grups, grups_a):
        file=open(f'{i}.txt' , 'r')
        dm=data.month
        x = file.readlines()
        d=0
        for mes in x:
            fin=[]
            for a in mes.split(' '):
                try:
                    a= int(a)
                    fin.append(a)
                except:
                    pass
            print(fin)
            f= sum(fin)
            ver= list(range(3, 26, 2))
            sheet[ga+str(ver[d])].alignment=alignment
            sheet[ga+str(ver[d])]= f
            d+=1
    
grups=['Трицепс', 'Бицепс', 'Грудь', 'Предплечья', 'Спина', 'Ноги', 'Икры']
res_grups={1:'Трицепс', 2:'Бицепс', 3:'Грудь', 4: 'Предплечья', 5:'Спина',6: 'Ноги', 7:'Икры'}
days=range(1,32)
months={'B':'Январь','C':'Февраль','D':'Март','E':'Апрель','F':'Май',
        'G':'Июнь','H':'Июль','I':'Август','J':'Сентябрь',
        'K':'Октябрь','L':'Ноябрь','M':'Декабрь'}
grups_a='NOPQRST'

data=datetime.date.today()
dm=data.month
dd=data.day

window = Tk()
window.title("Программа тренировки")
window.geometry('400x250')

file=open('Data.txt' , 'a')
file.write(f'\n{dm} {dd}')
file.close()
    
trc= Label(window, text=grups[0], font=("Arial Bold", 14) )  
trc.grid(column=0, row=0)
txt1 = Entry(window,width=10)  
txt1.grid(column=1, row=0)

bi= Label(window, text=grups[1], font=("Arial Bold", 14) )  
bi.grid(column=0, row=1)
txt2 = Entry(window,width=10)  
txt2.grid(column=1, row=1)

gr= Label(window, text=grups[2], font=("Arial Bold", 14) )  
gr.grid(column=0, row=2)
txt3 = Entry(window,width=10)  
txt3.grid(column=1, row=2)

pred= Label(window, text=grups[3], font=("Arial Bold", 14) )  
pred.grid(column=0, row=3)
txt4 = Entry(window,width=10)  
txt4.grid(column=1, row=3)

spi= Label(window, text=grups[4], font=("Arial Bold", 14) )  
spi.grid(column=0, row=4)
txt5 = Entry(window,width=10)  
txt5.grid(column=1, row=4)

nog= Label(window, text=grups[5], font=("Arial Bold", 14) )  
nog.grid(column=0, row=5)
txt6 = Entry(window,width=10)  
txt6.grid(column=1, row=5)

ikr= Label(window, text=grups[6], font=("Arial Bold", 14) )  
ikr.grid(column=0, row=6)
txt7 = Entry(window,width=10)  
txt7.grid(column=1, row=6)


btn2 = Button(window, text="Добавить значения", command=clicked2)
btn2.grid(column=3, row=0)


vse_data=[]
file=open('Data.txt' , 'r')
for i in file.readlines():
    i=i.rstrip('\n')
    vse_data.append(i)
vse_data=list(set(vse_data))

wb = openpyxl.Workbook()
wb.create_sheet(title = 'Первый лист', index = 0)
sheet = wb['Первый лист']

font = Font(name='Arial', size=14, italic=False, color='000000')
fill = PatternFill("solid", fgColor="DDDDDD")
fill_otmetka=PatternFill("solid", fgColor="4daffa")
fill_grups=PatternFill("solid", fgColor="007fff")
alignment = Alignment(horizontal="center", vertical="center")

osnova()
otmetka()

file.close()
window.mainloop()
wb.save('example.xlsx')
